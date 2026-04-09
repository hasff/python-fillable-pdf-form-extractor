from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
current_dir = Path(__file__).parent

INPUT_DIR = current_dir / "input"
OUTPUT_DIR = current_dir / "output"

OUTPUT_DIR.mkdir(exist_ok=True)


# ─────────────────────────────────────────────
# INSPECTION FUNCTIONS
# ─────────────────────────────────────────────
def get_doc_dimensions(fileName):
    dimensions = []

    with fitz.open(fileName) as pdf:
        page = pdf[0]
        dimensions = (page.rect.width, page.rect.height)

    return dimensions

def draw_boxes(input_pdf, output_pdf, pages_rects, color=(1, 0, 0), fill=None, width=0.5):
    doc = fitz.open(input_pdf)

    for page_num, page in enumerate(doc):
        page_rects = pages_rects[page_num]

        for rect_coords in page_rects:
            rect = fitz.Rect(rect_coords[0:4])

            page.draw_rect(
                rect,
                fill=fill,
                color=color,
                width=width
            )

    doc.save(output_pdf)

def draw_section_areas(fileName, output_file, sections):
    draw_boxes(
        input_pdf=fileName,
        output_pdf=output_file,
        pages_rects=[sections],
        width=1.5
    )

def inspect_form_sections(fileName):
    width, height = get_doc_dimensions(fileName)

    x1 = width * 0.08
    x2 = width * 0.92

    sections = [
        (x1, height * 0.14, x2, height * 0.34),  # 1. Personal Information
        (x1, height * 0.37, x2, height * 0.46),  # 2. Insurance Information
        (x1, height * 0.49, x2, height * 0.66),  # 3. Medical History
        (x1, height * 0.7, x2, height * 0.78),  # 4. Emergency Contact 
        (x1, height * 0.81, x2, height * 0.93),  # 5. Consent and Signature
    ]

    draw_section_areas(fileName=fileName, output_file=OUTPUT_DIR / "boxes.pdf", sections= sections)

    sections_names = [
        "Personal Information", "Insurance Information", "Medical History", "Emergency Contact", "Consent and Signature"
    ]
    print("sections = [")
    for i, section in enumerate(sections):
        print(f"    {section}, # {i+1}. {sections_names[i]}")
    print("]")   


# ─────────────────────────────────────────────
# EXTRACTION FUNCTIONS
# ─────────────────────────────────────────────
def get_all_widgets(fileName):
    widgets = []

    with fitz.open(fileName) as pdf:
        page = pdf[0]
        widgets = [w for w in page.widgets()]

    return widgets

def get_widgets_in_area(widgets, area_rect):
    return {
        w.field_name: w for w in widgets if fitz.Rect(w.rect).intersects(area_rect)
    }

def get_field_value_by_label(widgets, label):
    return next(
        (w.field_value for w in widgets.values() if label in w.field_label),
        None
    )

def extract_data_from_form(FILE, inspect=False):

    if inspect:
        inspect_form_sections(FILE)
        return None


    # Based on the visualization, we can define the sections as follows:
    sections = [
        (47.6220458984375, 117.86456787109375, 547.6535278320313, 286.2425219726563), # 1. Personal Information
        (47.6220458984375, 311.4992150878906, 547.6535278320313, 387.26929443359376), # 2. Insurance Information
        (47.6220458984375, 412.52598754882814, 547.6535278320313, 555.6472485351562), # 3. Medical History
        (47.6220458984375, 589.3228393554687, 547.6535278320313, 656.6740209960938), # 4. Emergency Contact
        (47.6220458984375, 681.9307141113281, 547.6535278320313, 782.9574865722657), # 5. Consent and Signature
    ]    

    all_widgets = get_all_widgets(FILE)

    personal_info_widgets       = get_widgets_in_area(all_widgets, fitz.Rect(sections[0]))
    insurance_info_widgets      = get_widgets_in_area(all_widgets, fitz.Rect(sections[1]))
    medical_history_widgets     = get_widgets_in_area(all_widgets, fitz.Rect(sections[2]))
    emergency_contact_widgets   = get_widgets_in_area(all_widgets, fitz.Rect(sections[3]))
    consent_signature_widgets   = get_widgets_in_area(all_widgets, fitz.Rect(sections[4]))

    # ─────────────────────────────────────────────
    # 1. Personal Information
    # ─────────────────────────────────────────────
    first_name_value    = get_field_value_by_label(personal_info_widgets, "First")
    last_name_value     = get_field_value_by_label(personal_info_widgets, "Last")
    date_of_birth_value = get_field_value_by_label(personal_info_widgets, "Date")
    gender_value        = get_field_value_by_label(personal_info_widgets, "Gender")
    national_id_value   = get_field_value_by_label(personal_info_widgets, "National")
    nhs_number_value    = get_field_value_by_label(personal_info_widgets, "NHS")
    address_value       = get_field_value_by_label(personal_info_widgets, "Home")
    city_value          = get_field_value_by_label(personal_info_widgets, "City")
    postal_code_value   = get_field_value_by_label(personal_info_widgets, "Postal")
    country_value       = get_field_value_by_label(personal_info_widgets, "Country")
    phone_number_value  = get_field_value_by_label(personal_info_widgets, "Phone")
    email_value         = get_field_value_by_label(personal_info_widgets, "Email")


    # ─────────────────────────────────────────────
    # 2. Insurance Information
    # ─────────────────────────────────────────────
    insurance_provider_value    = get_field_value_by_label(insurance_info_widgets, "Provider")
    policy_number_value         = get_field_value_by_label(insurance_info_widgets, "Policy")
    gp_name_value               = get_field_value_by_label(insurance_info_widgets, "Doctor Name")
    gp_phone_value              = get_field_value_by_label(insurance_info_widgets, "Phone Number")


    # ─────────────────────────────────────────────
    # 3. Medical History
    # ─────────────────────────────────────────────
    diabetes_value          = get_field_value_by_label(medical_history_widgets, "Diabetes")
    hypertension_value      = get_field_value_by_label(medical_history_widgets, "Hypertension")
    heart_disease_value     = get_field_value_by_label(medical_history_widgets, "Heart Disease")
    asthma_value            = get_field_value_by_label(medical_history_widgets, "Asthma")
    allergies_value         = get_field_value_by_label(medical_history_widgets, "Allergies")
    cancer_value            = get_field_value_by_label(medical_history_widgets, "Cancer")
    mental_health_value     = get_field_value_by_label(medical_history_widgets, "Mental Health")
    epilepsy_value          = get_field_value_by_label(medical_history_widgets, "Epilepsy")
    other_conditions_value  = get_field_value_by_label(medical_history_widgets, "Other")

    current_medications_value = get_field_value_by_label(medical_history_widgets, "Current Medications")
    known_allergies_value     = get_field_value_by_label(medical_history_widgets, "Known Allergies")

    # ────────────────────────────────────────────
    # 4. Emergency Contact
    # ────────────────────────────────────────────
    emergency_contact_name_value   = get_field_value_by_label(emergency_contact_widgets, "Full Name")
    relationship_value             = get_field_value_by_label(emergency_contact_widgets, "Relation")
    emergency_contact_phone_value   = get_field_value_by_label(emergency_contact_widgets, "Phone")
    emergency_contact_email_value   = get_field_value_by_label(emergency_contact_widgets, "Email")


    # ────────────────────────────────────────────
    # 5. Consent and Signature
    # ────────────────────────────────────────────    
    consent_1_value   = get_field_value_by_label(consent_signature_widgets, "I consent to the med")
    consent_2_value   = get_field_value_by_label(consent_signature_widgets, "I consent to the pro")
    consent_3_value   = get_field_value_by_label(consent_signature_widgets, "I consent to sha")
    consent_date_value    = get_field_value_by_label(consent_signature_widgets, "Date")


    return {
    "personal_info": {
            "First Name":     first_name_value,
            "Last Name":      last_name_value,
            "Date of Birth":  date_of_birth_value,
            "Gender":         gender_value,
            "National ID":    national_id_value,
            "NHS Number":     nhs_number_value,
            "Address":        address_value,
            "City":           city_value,
            "Postal Code":    postal_code_value,
            "Country":        country_value,
            "Phone":          phone_number_value,
            "Email":          email_value,
        },
        "insurance_info": {
            "Insurance Provider": insurance_provider_value,
            "Policy Number":      policy_number_value,
            "GP Name":            gp_name_value,
            "GP Phone":           gp_phone_value,
        },
        "medical_history": {
            "Diabetes":            diabetes_value,
            "Hypertension":        hypertension_value,
            "Heart Disease":       heart_disease_value,
            "Asthma":              asthma_value,
            "Allergies":           allergies_value,
            "Cancer":              cancer_value,
            "Mental Health":       mental_health_value,
            "Epilepsy":            epilepsy_value,
            "Other Conditions":    other_conditions_value,
            "Current Medications": current_medications_value,
            "Known Allergies":     known_allergies_value,
        },
        "emergency_contact": {
            "Full Name":     emergency_contact_name_value,
            "Relationship":  relationship_value,
            "Phone":         emergency_contact_phone_value,
            "Email":         emergency_contact_email_value,
        },
        "consent_signature": {
            "Medical Treatment Consent": consent_1_value,
            "Processing Consent":        consent_2_value,
            "Data Sharing Consent":      consent_3_value,
            "Date":                      consent_date_value,
        },
    }


# ─────────────────────────────────────────────
# EXCEL GENERATION FUNCTIONS
# ─────────────────────────────────────────────
def generate_excel(output_path, data: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Admission"

    # ── Styles ──────────────────────────────────────
    TEAL        = "FF008080"
    LIGHT_TEAL  = "FFE0F2F1"
    WHITE       = "FFFFFFFF"
    DARK_GREY   = "FF424242"

    header_font    = Font(name="Arial", bold=True, color=WHITE, size=11)
    label_font     = Font(name="Arial", bold=True, color=DARK_GREY, size=10)
    value_font     = Font(name="Arial", color=DARK_GREY, size=10)
    header_fill    = PatternFill("solid", start_color=TEAL)
    label_fill     = PatternFill("solid", start_color=LIGHT_TEAL)

    thin_border = Border(
        left=Side(style="thin", color="FFB2DFDB"),
        right=Side(style="thin", color="FFB2DFDB"),
        top=Side(style="thin", color="FFB2DFDB"),
        bottom=Side(style="thin", color="FFB2DFDB"),
    )

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def write_section_header(row, title):
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value     = title
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border
        ws[f"B{row}"].border = thin_border
        return row + 1

    def write_row(row, label, value):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=(value if value != "Off" else "No") or "—")
        lc.font      = label_font
        lc.fill      = label_fill
        lc.alignment = left
        lc.border    = thin_border
        vc.font      = value_font
        vc.fill      = PatternFill("solid", start_color=WHITE)
        vc.alignment = left
        vc.border    = thin_border
        ws.row_dimensions[row].height = 18
        return row + 1

    # ── Column widths ────────────────────────────────
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100

    # ── Sections ─────────────────────────────────────
    r = 1
    r = write_section_header(r, "1. Personal Information")
    for label, value in data["personal_info"].items():
        r = write_row(r, label, value)

    r += 1
    r = write_section_header(r, "2. Insurance Information")
    for label, value in data["insurance_info"].items():
        r = write_row(r, label, value)

    r += 1
    r = write_section_header(r, "3. Medical History")
    for label, value in data["medical_history"].items():
        r = write_row(r, label, value)

    r += 1
    r = write_section_header(r, "4. Emergency Contact")
    for label, value in data["emergency_contact"].items():
        r = write_row(r, label, value)

    r += 1
    r = write_section_header(r, "5. Consent and Signature")
    for label, value in data["consent_signature"].items():
        r = write_row(r, label, value)

    wb.save(output_path)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":

    FILE = INPUT_DIR / "patient_admission_form_john_doe.pdf"
    data = extract_data_from_form(FILE, inspect=False)

    if data:
        generate_excel(OUTPUT_DIR / "extracted_data.xlsx", data)
